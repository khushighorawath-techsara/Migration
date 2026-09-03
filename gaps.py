"""Are the remaining time mismatches a systematic offset, or just noise?

A tight cluster around 570 min means those rows are ET with no marker
(ET->IST is +9h30). A cluster at 720 means an AM/PM error in the sheet.
Scatter means the data genuinely does not line up and no rule will fix it.
"""
import openpyxl
from collections import Counter
ws = openpyxl.load_workbook('v5_final.xlsx').active
h  = [c.value for c in ws[1]]
G, V = h.index('Time gap (min)'), h.index('VERDICT')

gaps = [r[G] for r in ws.iter_rows(min_row=2, values_only=True)
        if isinstance(r[G], int) and r[V] and 'MISMATCH' in str(r[V])]
print(f"{len(gaps)} row(s) with a measurable gap and a mismatch verdict\n")

buckets = Counter()
for g in gaps:
    if   g <= 60:  buckets['under 1h']       += 1
    elif g <= 150: buckets['1-2.5h']         += 1
    elif 540 <= g <= 600: buckets['~9.5h  <- ET not marked'] += 1
    elif 690 <= g <= 750: buckets['~12h   <- AM/PM error']   += 1
    elif g <= 400: buckets['2.5-6.5h']       += 1
    else:          buckets['other / large']  += 1
for k, v in buckets.most_common():
    print(f"  {v:4d}  {k}")

print("\n  exact gap values seen (top 12):")
for k, v in Counter(gaps).most_common(12):
    print(f"    {v:3d} x  {k//60}h{k%60:02d}  ({k} min)")
