#!/usr/bin/env python3
"""
Finds the S3 meeting id for each row of a spreadsheet of sessions whose meeting
id is unknown.

READ-ONLY. Lists S3 and writes one local .xlsx. Nothing in S3, Zoom or Salesforce
is created, changed or deleted.

WHY CANDIDATE NAME IS THE PRIMARY KEY, NOT DATE+TIME
  The obvious approach is to match on date and time, since S3 stores a
  Time-H-MM-AM-IST folder per session. The source data does not support it:

    * Timezones are mixed and mostly unmarked. Some rows say IST, some say ET,
      most say nothing. ET and IST are 9.5 hours apart, so matching an unmarked
      time is worse than not matching at all -- it produces confident wrong
      answers.
    * Several rows have no AM/PM: "8.30 - 9 IST" could be 08:30 or 20:30.
    * Some dates are text ("Today", "13/08/2026") or shifted into the wrong
      column entirely.

  The candidate name is present on every row and appears verbatim in the S3
  path, so that is the anchor. Date filters the result, and time only ranks
  between several matches. It is never allowed to exclude one.

CONFIDENCE TIERS -- every row gets an answer, even if that answer is "look here"
  A  exactly one session for this candidate within the date window   auto-accept
  B  several within the window -- ranked by time, all returned       likely
  C  candidate found but not on that date -- all their sessions      review
  D  only a fuzzy or partial name match                             review
  E  nothing found                                                  absent

  Tiers C and D still return meeting ids. A row is only reported as absent when
  the candidate does not appear anywhere in the bucket.

USAGE
  python3 find_missing_meeting_ids.py --xlsx Missing_MeetingIDs.xlsx
  python3 find_missing_meeting_ids.py --xlsx in.xlsx --out results.xlsx --day-window 2
"""

import argparse
import datetime
import os
import re
import sys
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher

try:
    import boto3
    import openpyxl
except ImportError as exc:
    print(f"Missing dependency: {exc}\nRun: pip install boto3 openpyxl", file=sys.stderr)
    sys.exit(1)

BUCKET     = os.environ.get("MIGRATION_BUCKET", "zoom-automation-bucket")
AWS_REGION = os.environ.get("AWS_DEFAULT_REGION", "us-east-1")

DATE_RE  = re.compile(r"^\d{4}-\d{2}-\d{2}$")
YEAR_RE  = re.compile(r"^\d{4}$")
MID_RE   = re.compile(r"^\d{9,11}$")
TIME_RE  = re.compile(r"^Time-(\d{1,2})-(\d{2})-(AM|PM)-", re.I)
TIME_RE2 = re.compile(r"^Time-(\d{1,2})-(AM|PM)-", re.I)

# Rows whose Date cell is unusable, resolved from their position in the
# date-ordered sheet. Row 110's date cell holds a candidate name (a column
# shift); rows 48-51 say "Today". Both were pinned by looking at the rows
# either side, not guessed.
DATE_OVERRIDES = {
    48: datetime.date(2026, 8, 4),   # sits between 08-03 and 08-05
    49: datetime.date(2026, 8, 4),
    50: datetime.date(2026, 8, 4),
    51: datetime.date(2026, 8, 4),
    110: datetime.date(2026, 8, 13), # column shift; neighbours are all 08-13
}

# Tokens that carry no identifying information and appear on one side only.
HOST_MATCH_BONUS = float(os.environ.get("HOST_MATCH_BONUS", "0.60"))

NOISE_TOKENS = {"na", "n", "a", "ms", "mr", "mrs", "dr", "not", "applicable"}

s3 = boto3.client("s3", region_name=AWS_REGION)


def log(m=""):
    print(m, flush=True)


# ══════════════════════════════════════════════════════════════════════════════
#  Normalisation -- the same function must run over BOTH sides
# ══════════════════════════════════════════════════════════════════════════════

def norm_name(raw) -> str:
    """
    Reduce a person's name to comparable tokens.

    The two sides format names differently, so both are collapsed to the same
    shape rather than trying to guess one from the other:

        "Abdu Raziq N/A Hidayathulla"      -> "abdu raziq hidayathulla"
        "Abdu_Raziq_N_A_Hidayathulla"      -> "abdu raziq hidayathulla"
        "Sindooja Gajam, Ms."              -> "sindooja gajam"
        "Dhruvi Rajubhai . Ramani"         -> "dhruvi rajubhai ramani"

    "N/A" appears as a literal middle name on both sides and is dropped, along
    with honorifics -- they are noise that would otherwise block an exact match.
    """
    s = unicodedata.normalize("NFKD", str(raw or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace("_", " ").replace("/", " ").replace(".", " ").replace(",", " ")
    s = re.sub(r"[^A-Za-z0-9 ]+", " ", s).lower()
    toks = [t for t in s.split() if t and t not in NOISE_TOKENS]
    return " ".join(toks)


def name_tokens(raw) -> set:
    return set(norm_name(raw).split())


def names_match(a: str, b: str):
    """
    Returns (matched, how, score).

    Three levels, strongest first. Token-subset matters because the spreadsheet
    often carries only a first name ("Sindooja", "Bala Kishor") where S3 has the
    full one ("Sindooja_Gajam", "Bala_Kishore_Gonga"). Requiring an exact match
    would drop those rows entirely.
    """
    na, nb = norm_name(a), norm_name(b)
    if not na or not nb:
        return False, "", 0.0
    if na == nb:
        return True, "exact", 1.0

    ta, tb = set(na.split()), set(nb.split())
    if ta and tb and (ta <= tb or tb <= ta):
        # every token of the shorter name is present in the longer one
        return True, "subset", 0.9

    # Truncated or differently-spelled tokens: "Bala Kishor" vs
    # "Bala_Kishore_Gonga", "Jagadeeshh" vs "Jagadeesh". Every token of the
    # shorter name must PREFIX a distinct token of the longer one, and at least
    # one of them must be 5+ characters -- without that length floor a lone
    # "Sai" would prefix half the bucket.
    short, long_ = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
    if short and max((len(t) for t in short), default=0) >= 5:
        remaining = set(long_)
        ok_all = True
        for t in short:
            hit = next((u for u in remaining if u.startswith(t) or t.startswith(u)), None)
            if hit is None:
                ok_all = False
                break
            remaining.discard(hit)
        if ok_all:
            return True, "prefix", 0.8

    ratio = SequenceMatcher(None, na, nb).ratio()
    if ratio >= 0.86:
        return True, f"fuzzy:{ratio:.2f}", ratio
    return False, "", ratio


# ══════════════════════════════════════════════════════════════════════════════
#  Source spreadsheet
# ══════════════════════════════════════════════════════════════════════════════

def parse_source_date(value, record_no):
    """A real date, a dd/mm/yyyy string, or a pinned override. Else None."""
    if record_no in DATE_OVERRIDES:
        return DATE_OVERRIDES[record_no], "pinned from sheet order"
    if isinstance(value, datetime.datetime):
        return value.date(), ""
    if isinstance(value, datetime.date):
        return value, ""

    txt = str(value or "").strip()
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", txt)
    if m:                                    # dd/mm/yyyy -- day first
        d, mo, y = (int(x) for x in m.groups())
        try:
            return datetime.date(y, mo, d), "parsed dd/mm/yyyy"
        except ValueError:
            pass
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", txt)
    if m:                                    # "18 August 2026 (EST)"
        try:
            return datetime.datetime.strptime(
                f"{m.group(1)} {m.group(2)[:3]} {m.group(3)}", "%d %b %Y").date(), "parsed long form"
        except ValueError:
            pass
    return None, f"unusable date {txt!r}"


# US daylight saving 2026: 8 Mar - 1 Nov. Every date in this dataset falls
# inside it, but the boundary is computed rather than assumed so the script
# stays correct if it is reused on winter data.
def _us_dst(d: datetime.date) -> bool:
    return datetime.date(2026, 3, 8) <= d <= datetime.date(2026, 11, 1) if d.year == 2026 \
        else 3 <= d.month <= 10

# Minutes to ADD to a local time to reach IST. S3 folders are IST, so
# everything is normalised to that.
def _offset_to_ist(tz: str, d: datetime.date) -> int:
    tz = (tz or "").upper()
    dst = _us_dst(d) if d else True
    if tz in ("IST", ""):          return 0
    if tz in ("ET", "EASTERN"):    return 570 if dst else 630     # EDT 9.5h / EST 10.5h
    if tz == "EDT":                return 570
    if tz == "EST":                return 630
    if tz in ("CT", "CENTRAL"):    return 630 if dst else 690
    if tz == "CDT":                return 630
    if tz == "CST":                return 690
    if tz in ("PT", "PACIFIC"):    return 750 if dst else 810
    if tz == "PDT":                return 750
    if tz == "PST":                return 810
    if tz in ("UTC", "GMT"):       return 330
    return 0                                                      # unknown -> treat as IST


TZ_RE = re.compile(r"\b(IST|EDT|EST|ET|CDT|CST|CT|PDT|PST|PT|UTC|GMT)\b", re.I)


def parse_source_time(value, source_date=None):
    """
    Every plausible reading of the time cell, converted to IST.

    Returns a list of (minutes_past_midnight_IST, day_offset, note). A list,
    not a single value, because two things in this data are genuinely
    ambiguous and guessing either one would be worse than carrying both:

      1. NO AM/PM -- "8.30 - 9 IST" is either 08:30 or 20:30. Rather than pick,
         BOTH are returned and the ranker keeps whichever actually lines up
         with a real session. The S3 data disambiguates instead of us.

      2. NO TIMEZONE -- most rows carry no marker. Those are read as IST,
         because S3 stores IST and every row that IS marked says IST far more
         often than anything else. Rows explicitly marked ET/EST/PT are
         converted properly.

    day_offset matters: 9 PM ET is 6:30 AM IST the FOLLOWING day, so the
    conversion can move the calendar date. Ignoring that would look for the
    session under the wrong date entirely.
    """
    txt = str(value or "").strip()
    if not txt:
        return []

    tzm = TZ_RE.search(txt)
    tz  = tzm.group(1).upper() if tzm else ""
    off = _offset_to_ist(tz, source_date)
    note = f"{tz or 'no tz, read as IST'}"
    if off:
        note += f", +{off // 60}h{off % 60:02d} to IST"

    def to_ist(h24, mi):
        total = h24 * 60 + mi + off
        return total % 1440, total // 1440

    # "10:00 PM", "08:30 PM", "1:00 AM"
    m = re.match(r"^\s*(\d{1,2})[:.](\d{2})\s*(AM|PM)", txt, re.I)
    if m:
        h, mi, ap = int(m.group(1)), int(m.group(2)), m.group(3).upper()
        if ap == "PM" and h != 12: h += 12
        if ap == "AM" and h == 12: h = 0
        t, dz = to_ist(h, mi)
        return [(t, dz, note)]

    # "1 PM", "9 AM"
    m = re.match(r"^\s*(\d{1,2})\s*(AM|PM)", txt, re.I)
    if m:
        h, ap = int(m.group(1)), m.group(2).upper()
        if ap == "PM" and h != 12: h += 12
        if ap == "AM" and h == 12: h = 0
        t, dz = to_ist(h, 0)
        return [(t, dz, note)]

    # "8.30 - 9 IST", "11.30 - 12 IST", "3 - 3.30 IST", "1:00 ET - 1:30 ET"
    # -- a clock reading with no AM/PM. Keep both readings.
    m = re.match(r"^\s*(\d{1,2})(?:[:.](\d{2}))?\s*(?:[-–—]|\bto\b|[A-Za-z])", txt)
    if m:
        h  = int(m.group(1))
        mi = int(m.group(2) or 0)
        if 0 <= h <= 23 and 0 <= mi <= 59:
            out = []
            am_h = 0 if h == 12 else h
            pm_h = 12 if h == 12 else (h + 12 if h < 12 else h)
            for hh, lbl in ((am_h, "read as AM"), (pm_h, "read as PM")):
                if hh > 23:
                    continue
                t, dz = to_ist(hh, mi)
                if not any(t == o[0] and dz == o[1] for o in out):
                    out.append((t, dz, f"{note}; no AM/PM, {lbl}"))
            return out
    return []


def read_source(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in r):
            continue
        rec = int(r[0]) if isinstance(r[0], (int, float)) else None
        d, dnote = parse_source_date(r[1], rec)
        out.append({
            "record": rec, "raw": list(r), "date": d, "date_note": dnote,
            "times": parse_source_time(r[2], d),
            "candidate": r[3], "trainer": r[4], "proxy": r[5],
        })
    return header, out


# ══════════════════════════════════════════════════════════════════════════════
#  S3 index
# ══════════════════════════════════════════════════════════════════════════════

def parse_session_key(key):
    """
    Pull one session's identity out of an object key, WITHOUT hardcoding any
    department's layout.

    The bucket contains at least four shapes -- Interview-Success old and new,
    a legacy variant with the meeting id early, and the generic department
    layout. Rather than encode all of them, this anchors on the 4-digit year,
    because the relationship around it holds in every one:

        {...}/{Host}/{YYYY}/{Month}/{Candidate}/{...}/{MeetingID}/{Media}/file

    host is always immediately before the year, candidate two after it. A
    positional parser would need updating every time a layout changes; this does
    not.
    """
    parts = key.split("/")
    if len(parts) < 6:
        return None
    year_idx = next((i for i, p in enumerate(parts) if i > 0 and YEAR_RE.match(p)), None)
    if year_idx is None or year_idx < 1 or year_idx + 2 >= len(parts):
        return None

    mid = next((p for p in parts[year_idx:] if MID_RE.match(p)), None)
    if not mid:
        return None

    date = next((p for p in parts if DATE_RE.match(p)), None)
    tmin = None
    for p in parts:
        m = TIME_RE.match(p) or TIME_RE2.match(p)
        if m:
            g = m.groups()
            h = int(g[0]); mi = int(g[1]) if len(g) == 3 else 0
            ap = (g[-1]).upper()
            if ap == "PM" and h != 12: h += 12
            if ap == "AM" and h == 12: h = 0
            tmin = h * 60 + mi
            break

    return {
        "department": parts[0],
        "host":       parts[year_idx - 1],
        "candidate":  parts[year_idx + 2],
        "date":       datetime.date.fromisoformat(date) if date else None,
        "time_min":   tmin,
        "meeting_id": mid,
        "prefix":     "/".join(parts[: parts.index(mid) + 1]) + "/",
    }


def build_index():
    """One pass over the bucket -> one row per distinct session."""
    sessions, scanned = {}, 0
    paginator = s3.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=BUCKET):
        for obj in page.get("Contents", []):
            scanned += 1
            key = obj["Key"]
            if key.startswith(("shortlinks/", "zoom-fail/", "temp/", "ec2-jobs/")):
                continue
            info = parse_session_key(key)
            if info and info["meeting_id"] not in sessions:
                sessions[info["meeting_id"]] = info
        if scanned % 100000 < 1200:
            log(f"    ...scanned {scanned} objects, {len(sessions)} session(s)")
    log(f"  {scanned} objects -> {len(sessions)} distinct session(s)")
    return list(sessions.values())


# ══════════════════════════════════════════════════════════════════════════════
#  Matching
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
#  Real meeting start times, from Zoom
# ══════════════════════════════════════════════════════════════════════════════
#
# WHY THIS EXISTS
#   Only the generic department layout carries a Time-H-MM-AM-IST folder.
#   Interview-Success uses {Company}/{Date}/{Round}/ instead, so roughly two
#   thirds of sessions have NO clock anywhere in their S3 path. Time therefore
#   could not rank them, and a human cannot verify them by time either.
#
#   That matters most exactly where it is scarcest: when one candidate has
#   several sessions on the same day with different hosts, the clock is the
#   only thing that separates them.
#
#   Zoom knows the true start_time for every meeting id. Fetching it turns time
#   from a missing signal into a verifiable column on every row.
#
# Optional by design -- without credentials the script still runs exactly as
# before, using whatever Time- folders happen to exist.

IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))
_zoom_cache = {}


def _zoom_token():
    """Authenticate by reusing the processor Lambda's OWN credential code.

    A standalone loader has to guess what the secret's JSON keys are called,
    and guessing wrong fails with 'no Zoom credentials' -- which reads like a
    permissions problem when it is actually a key-name problem. The Lambda
    already knows the shape and authenticates with it in production every day,
    so its functions are called rather than reimplemented.

    Needs lambda_function.py beside this script; without it, time verification
    is simply skipped and the run continues.
    """
    path = os.environ.get("PROCESSOR_LAMBDA_PATH", "lambda_function.py")
    if not os.path.exists(path):
        log(f"  ({path} not found -- skipping Zoom time lookup)")
        return None
    try:
        import importlib.util
        os.environ.setdefault("ZOOM_SECRET_NAME",
                              os.environ.get("ZOOM_SECRET_NAME", "zoom/general-oauth"))
        os.environ.setdefault("S3_BUCKET_NAME", BUCKET)
        spec = importlib.util.spec_from_file_location("_proc", path)
        proc = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(proc)
        tok = proc.get_s2s_access_token(proc.get_zoom_secret())
        log(f"  authenticated via Secrets Manager "
            f"({getattr(proc, 'ZOOM_SECRET_NAME', 'zoom secret')})")
        return tok
    except Exception as exc:
        log(f"  (Zoom auth failed: {exc})")
        return None


def zoom_start_ist(token, meeting_id):
    """(minutes past midnight IST, date) for a meeting, or (None, None).

    Zoom returns UTC; S3 and the source sheet are IST, so it is converted here
    and the DATE is returned too -- a UTC evening meeting is the next day in
    IST, and comparing against the wrong date would look like a mismatch.
    """
    if not token or not meeting_id:
        return None, None
    if meeting_id in _zoom_cache:
        return _zoom_cache[meeting_id]
    try:
        import requests
        r = requests.get(f"https://api.zoom.us/v2/past_meetings/{meeting_id}",
                         headers={"Authorization": f"Bearer {token}"}, timeout=30)
        if r.status_code != 200:
            _zoom_cache[meeting_id] = (None, None)
            return None, None
        st = r.json().get("start_time")
        if not st:
            _zoom_cache[meeting_id] = (None, None)
            return None, None
        dt = datetime.datetime.fromisoformat(st.replace("Z", "+00:00")).astimezone(IST)
        out = (dt.hour * 60 + dt.minute, dt.date())
        _zoom_cache[meeting_id] = out
        return out
    except Exception:
        _zoom_cache[meeting_id] = (None, None)
        return None, None


def match_row(row, sessions, by_cand_token, day_window):
    """Return (tier, reason, [ranked matches])."""
    cand = row["candidate"]
    if not cand:
        return "E", "no candidate name in the source row", []

    # Narrow by shared token first -- comparing against every session for every
    # row is O(rows x sessions) and needlessly slow.
    toks = name_tokens(cand)
    pool_ids, pool = set(), []
    for t in toks:
        for s in by_cand_token.get(t, ()):
            if id(s) not in pool_ids:
                pool_ids.add(id(s)); pool.append(s)
    if not pool:
        pool = sessions                       # fall back to a full scan

    hits = []
    for s in pool:
        ok, how, score = names_match(cand, s["candidate"])
        if ok:
            hits.append((s, how, score))
    if not hits:
        return "E", "candidate name not found anywhere in the bucket", []

    # The Zoom host may be EITHER the Trainer/Interviewer OR the Proxy Person --
    # which of the two it is, is not known per row. So both are tested and
    # whichever matches is reported back, turning an unknown into an answer.
    #
    # Still a TIEBREAKER ONLY, never a filter. S3's folder is the host; if
    # neither name matches it may simply mean a third person hosted, which is
    # not grounds for rejecting an otherwise good candidate+date match.
    def host_role(s):
        """Returns ('trainer'|'proxy'|'', bonus)."""
        for label, val in (("trainer", row.get("trainer")), ("proxy", row.get("proxy"))):
            if not val:
                continue
            ok, _, _ = names_match(val, s["host"])
            if ok:
                # Deliberately large. A trainer or proxy name matching the S3
                # host folder is the STRONGEST evidence available -- it is a
                # second independent person agreeing, not a spelling variant.
                #
                # At 0.18 it was being swamped: two sessions for the same
                # candidate on the same day, one hosted by the named trainer
                # with the candidate's name spelled in a different order, the
                # other hosted by a stranger with the name spelled exactly.
                # The name gap (~0.3) beat the host bonus and the stranger's
                # session won. Ranking spelling above identity is backwards.
                return label, HOST_MATCH_BONUS
        return "", 0.0

    def trainer_bonus(s):
        return host_role(s)[1]

    sdate = row["date"]
    times = row.get("times") or []

    # A timezone conversion can push the session onto the next calendar day
    # (9 PM ET is 6:30 AM IST tomorrow), so the effective date to search
    # around is the source date PLUS whatever day_offset the conversion
    # produced. Every reading contributes its own effective date.
    eff_dates = set()
    if sdate:
        eff_dates.add(sdate)
        for _t, dz, _n in times:
            if dz:
                eff_dates.add(sdate + datetime.timedelta(days=dz))

    if eff_dates:
        in_window = [h for h in hits
                     if h[0]["date"] and any(
                         abs((h[0]["date"] - ed).days) <= day_window for ed in eff_dates)]
    else:
        in_window = []

    def time_bonus(s):
        """Best score across every plausible reading of the time cell.

        Where a cell had no AM/PM, both readings were kept -- the one that
        actually lines up with a real session wins here. That is the S3 data
        resolving the ambiguity rather than us guessing at it."""
        if not times or s["time_min"] is None:
            return 0.0
        best = 0.0
        for t, _dz, _n in times:
            diff = min(abs(t - s["time_min"]), 1440 - abs(t - s["time_min"]))
            best = max(best, max(0.0, 0.25 - diff / 480.0))
        return best

    def rank(h):
        s, how, score = h
        r = score + trainer_bonus(s) + time_bonus(s)
        if eff_dates and s["date"]:
            r += max(0.0, 0.30 - 0.10 * min(abs((s["date"] - ed).days) for ed in eff_dates))
        return -r

    if in_window:
        in_window.sort(key=rank)
        exact = [h for h in in_window if h[1] in ("exact", "subset")]
        if len(in_window) == 1:
            return "A", f"one session for this candidate within +/-{day_window}d", in_window
        if exact and len(exact) == 1:
            return "A", f"one strong name match within +/-{day_window}d", exact
        return "B", f"{len(in_window)} sessions within +/-{day_window}d, ranked", in_window

    hits.sort(key=rank)
    strong = [h for h in hits if h[1] in ("exact", "subset")]
    if strong:
        note = "date unusable" if not sdate else "no session on that date"
        return "C", f"candidate found but {note}; all their sessions returned", strong[:8]
    return "D", "only a fuzzy name match -- needs a human check", hits[:5]


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--xlsx", required=True, help="Source spreadsheet.")
    p.add_argument("--out", default="meeting_id_results.xlsx", help="Output file.")
    p.add_argument("--no-zoom", action="store_true",
                   help="Skip the Zoom start-time lookup. Without it, time can "
                        "only be checked on the ~1/3 of sessions whose S3 path "
                        "contains a Time- folder.")
    p.add_argument("--time-tolerance", type=int, default=45,
                   help="Minutes. A pick whose real start differs by more than "
                        "this is flagged for review (default 45).")
    p.add_argument("--day-window", type=int, default=1,
                   help="Days either side of the source date to accept (default 1). "
                        "One day absorbs the timezone problem: a 2 AM IST session "
                        "lands on the next calendar day in S3.")
    args = p.parse_args()

    log(f"=== Meeting-id lookup — READ ONLY ===")
    log(f"Bucket: {BUCKET}   date window: +/-{args.day_window} day(s)\n")

    log(f"[1/4] Reading {args.xlsx} ...")
    header, rows = read_source(args.xlsx)
    log(f"  {len(rows)} row(s)")
    log(f"  {sum(1 for r in rows if not r['date'])} row(s) still have no usable date")
    log(f"  {sum(1 for r in rows if not r['times'])} row(s) have no parseable time")
    log(f"  {sum(1 for r in rows if len(r['times']) > 1)} row(s) had no AM/PM — "
        f"both readings kept, S3 decides which")
    log(f"  {sum(1 for r in rows if any(d for _t, d, _n in r['times']))} row(s) shift "
        f"a calendar day after timezone conversion\n")

    log("[2/4] Indexing S3 ...")
    sessions = build_index()

    by_tok = defaultdict(list)
    for s in sessions:
        for t in name_tokens(s["candidate"]):
            by_tok[t].append(s)
    log(f"  indexed {len(by_tok)} distinct candidate token(s)\n")

    log("[3/4] Matching ...")
    results, tally = [], defaultdict(int)
    for r in rows:
        tier, reason, hits = match_row(r, sessions, by_tok, args.day_window)
        tally[tier] += 1
        best = hits[0][0] if hits else None
        who = ""
        if best:
            for label, val in (("Trainer/Interviewer", r.get("trainer")),
                               ("Proxy Person", r.get("proxy"))):
                if val:
                    ok, _, _ = names_match(val, best["host"])
                    if ok:
                        who = label
                        break
            if not who:
                who = "neither -- a third person hosted"
        alts = "; ".join(
            f"{h[0]['meeting_id']} ({h[0]['date']} {h[0]['department']})" for h in hits[1:6])
        results.append({
            "row": r, "tier": tier, "reason": reason, "hits": hits,
            "meeting_id": best["meeting_id"] if best else "",
            "prefix":     best["prefix"] if best else "",
            "department": best["department"] if best else "",
            "s3_candidate": best["candidate"] if best else "",
            "s3_host":      best["host"] if best else "",
            "s3_date":      str(best["date"]) if best and best["date"] else "",
            "how":  hits[0][1] if hits else "",
            "host_is": who,
            "tz_note": "; ".join(sorted({n for _t, _d, n in (r.get("times") or [])})),
            "alts": alts, "n_alts": max(0, len(hits) - 1),
        })

    # ── real start times ────────────────────────────────────────────────────
    if not args.no_zoom:
        log("\n  Fetching real start times from Zoom ...")
        tok = _zoom_token()
        if not tok:
            log("  no Zoom credentials -- falling back to Time- folders only\n")
        else:
            done = 0
            for res in results:
                if not res["meeting_id"]:
                    continue
                tmin, tdate = zoom_start_ist(tok, res["meeting_id"])
                res["zoom_time"] = tmin
                res["zoom_date"] = tdate
                done += 1
                if done % 40 == 0:
                    log(f"    ...{done} looked up")
            got = sum(1 for r in results if r.get("zoom_time") is not None)
            log(f"  got a real start time for {got} of {done} meeting(s)\n")

            # ── RE-PICK using real times ────────────────────────────────
            # The ranker chose before any real clock was known, using name,
            # date and whatever Time- folder happened to exist. Now that Zoom
            # has given a true start for every meeting, any row whose pick is
            # far from its scheduled slot is revisited: the alternatives are
            # looked up too, and if one lands closer it takes over.
            #
            # This is the opposite of loosening a threshold. It uses a NEW,
            # harder fact to overturn a weaker guess -- the same candidate on
            # the same day with two sessions is exactly the case the clock
            # exists to resolve.
            repicked = 0
            for res in results:
                row = res["row"]
                if not res["meeting_id"] or not row.get("times"):
                    continue
                zt = res.get("zoom_time")
                cur = (min(min(abs(t - zt), 1440 - abs(t - zt)) for t, _d, _n in row["times"])
                       if zt is not None else 10**6)
                if cur <= args.time_tolerance:
                    continue                      # already good

                best, best_gap = None, cur
                for s, how, score in res.get("hits", [])[:8]:
                    if s["meeting_id"] == res["meeting_id"]:
                        continue
                    at, ad = zoom_start_ist(tok, s["meeting_id"])
                    if at is None:
                        continue
                    g = min(min(abs(t - at), 1440 - abs(t - at)) for t, _d, _n in row["times"])
                    # Only overturn on a CLEARLY better fit, and only when the
                    # replacement is itself within tolerance. A marginal
                    # improvement is not evidence.
                    if g < best_gap and g <= args.time_tolerance:
                        best, best_gap = (s, how, at, ad), g

                if best:
                    s, how, at, ad = best
                    res.update({"meeting_id": s["meeting_id"], "prefix": s["prefix"],
                                "department": s["department"], "s3_candidate": s["candidate"],
                                "s3_host": s["host"],
                                "s3_date": str(s["date"]) if s["date"] else "",
                                "how": how, "zoom_time": at, "zoom_date": ad})
                    who = ""
                    for lbl, val in (("Trainer/Interviewer", row.get("trainer")),
                                     ("Proxy Person", row.get("proxy"))):
                        if val and names_match(val, s["host"])[0]:
                            who = lbl; break
                    res["host_is"] = who or "neither -- a third person hosted"
                    res["reason"] += "  [re-picked on real start time]"
                    repicked += 1
            if repicked:
                log(f"  re-picked {repicked} row(s) using the real start time")

    # ── compare against the source time ─────────────────────────────────────
    flagged = 0
    for res in results:
        row = res["row"]
        zt  = res.get("zoom_time")
        if zt is None or not row.get("times"):
            res["time_gap"] = ""
            res["time_ok"]  = "no clock available" if zt is None else ""
            continue
        best = min(min(abs(t - zt), 1440 - abs(t - zt)) for t, _d, _n in row["times"])
        res["time_gap"] = best
        if best <= 5:
            res["time_ok"] = "exact"
        elif best <= args.time_tolerance:
            res["time_ok"] = "close"
        else:
            res["time_ok"] = f"OFF BY {best//60}h{best%60:02d} — CHECK"
            flagged += 1
    if flagged:
        log(f"  {flagged} row(s) have a start time more than "
            f"{args.time_tolerance} min from the scheduled slot\n")

    # ── the strict verdict ──────────────────────────────────────────────────
    # The stated bar is date AND time AND candidate. Anything short of that is
    # a lead, not an answer -- so it is labelled as such rather than being
    # counted alongside the confirmed rows.
    import datetime as _dt
    verdict_tally = defaultdict(int)
    for res in results:
        row = res["row"]
        if not res["meeting_id"]:
            res["verdict"] = "NO MATCH"; verdict_tally["NO MATCH"] += 1; continue

        name_ok = res["how"] in ("exact", "subset")
        date_ok = bool(row["date"]) and str(res["s3_date"]) == row["date"].isoformat()
        # Prefer Zoom's real start; fall back to the S3 date when Zoom had none.
        if res.get("zoom_date"):
            date_ok = bool(row["date"]) and res["zoom_date"] == row["date"]
        gap = res.get("time_gap", "")
        time_ok = isinstance(gap, int) and gap <= args.time_tolerance
        time_unknown = not isinstance(gap, int)
        host_ok = bool(res["host_is"]) and "neither" not in res["host_is"]

        if name_ok and date_ok and time_ok:
            v = "VERIFIED" if host_ok else "VERIFIED (other host)"
        elif name_ok and date_ok and time_unknown:
            v = "NO TIME — unconfirmed"
        elif name_ok and date_ok:
            v = "TIME MISMATCH — check"
        elif name_ok and time_ok:
            v = "DATE MISMATCH — check"
        else:
            v = "WEAK — check"
        res["verdict"] = v
        verdict_tally[v] += 1

    log("\n=== Verdict (date + time + candidate all confirmed) ===")
    for k in ("VERIFIED", "VERIFIED (other host)", "NO TIME — unconfirmed",
              "TIME MISMATCH — check", "DATE MISMATCH — check", "WEAK — check",
              "NO MATCH"):
        if verdict_tally[k]:
            log(f"  {k:26s} {verdict_tally[k]:4d}")
    strong = verdict_tally["VERIFIED"] + verdict_tally["VERIFIED (other host)"]
    log(f"\n  {strong} row(s) meet the full bar")

    log("\n=== Result ===")
    labels = {"A": "A  confident, one match",
              "B": "B  several matches, ranked",
              "C": "C  candidate found, date did not line up",
              "D": "D  fuzzy name only, needs review",
              "E": "E  not found anywhere"}
    for t in "ABCDE":
        log(f"  {labels[t]:44s} {tally[t]:4d}")
    found = sum(tally[t] for t in "ABCD")
    log(f"\n  {found} of {len(rows)} row(s) have at least one meeting id "
        f"({100*found/max(1,len(rows)):.0f}%)")

    log(f"\n[4/4] Writing {args.out} ...")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Results"
    ws.append(list(header) + ["Meeting ID", "Confidence", "Why", "Department",
                              "S3 candidate", "S3 host", "Who hosted", "S3 date",
                              "Name match", "Time reading", "Actual start (IST)",
                              "Time gap (min)", "Time check", "VERDICT",
                              "Other candidates", "S3 prefix"])
    for r in results:
        ws.append(list(r["row"]["raw"]) + [
            r["meeting_id"], r["tier"], r["reason"], r["department"],
            r["s3_candidate"], r["s3_host"], r["host_is"], r["s3_date"],
            r["how"], r["tz_note"],
            (f"{r['zoom_time']//60:02d}:{r['zoom_time']%60:02d}"
             if r.get("zoom_time") is not None else ""),
            r.get("time_gap", ""), r.get("time_ok", ""), r.get("verdict", ""),
            r["alts"], r["prefix"]])
    for col, w in zip("ABCDEFGHIJKLMNOPQRSTUV",
                      [8,12,24,26,20,16,14,30,14,11,42,18,24,20,22,12,12,26,
                       16,12,24,22,22]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(args.out)

    log(f"  saved {os.path.abspath(args.out)}")
    log("\nTiers A and B are usable as-is. C and D still carry meeting ids but "
        "should be eyeballed.\nNothing in S3, Zoom or Salesforce was modified.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("\nInterrupted.")
        sys.exit(130)
