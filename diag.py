"""Show, for rows where the two runs disagree, what each one picked and why."""
import openpyxl, sys

def load(f):
    ws = openpyxl.load_workbook(f).active
    h = [c.value for c in ws[1]]
    return h, [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]

h1, a = load('meeting_id_results.xlsx')
h2, b = load('results_exact.xlsx')

def col(h, name): return h.index(name)
MI, CF, SD, SC, SH, WH, TR = (col(h1,'Meeting ID'), col(h1,'Confidence'),
    col(h1,'S3 date'), col(h1,'S3 candidate'), col(h1,'S3 host'),
    col(h1,'Who hosted'), col(h1,'Time reading'))
DATE, TIME, CAND, TRAINER, PROXY = 1, 2, 3, 4, 5

n = 0
for i, (r1, r2) in enumerate(zip(a, b)):
    if r1[MI] == r2[MI]:
        continue
    n += 1
    if n > 6:
        break
    print(f"--- sheet row {i+2}  |  source: {str(r1[DATE])[:10]}  {r1[TIME]}")
    print(f"    candidate={r1[CAND]!r}  trainer={r1[TRAINER]!r}  proxy={r1[PROXY]!r}")
    print(f"    time read as: {r1[TR]}")
    print(f"    +/-1day : {r1[MI]}  s3date={r1[SD]}  cand={r1[SC]}  host={r1[SH]}  [{r1[WH]}]  ({r1[CF]})")
    print(f"    exact   : {r2[MI]}  s3date={r2[SD]}  cand={r2[SC]}  host={r2[SH]}  [{r2[WH]}]  ({r2[CF]})")
    same = "SAME DATE as source" if str(r2[SD]) == str(r1[DATE])[:10] else "date differs"
    print(f"    -> exact pick is {same}\n")
