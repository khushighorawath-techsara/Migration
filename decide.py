"""Which run should we trust? Count how often date and host AGREE.

A pick where the S3 date equals the source date AND the S3 host is the named
trainer or proxy has two independent confirmations. That is the only metric
here that is not circular -- 'more host matches' can be gamed by weighting
host higher, but 'both agree' cannot.
"""
import openpyxl, datetime

def analyse(path):
    ws = openpyxl.load_workbook(path).active
    h  = [c.value for c in ws[1]]
    MI, SD, WH = h.index('Meeting ID'), h.index('S3 date'), h.index('Who hosted')
    both = date_only = host_only = neither = blank = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[MI]:
            blank += 1; continue
        src = r[1]
        srcd = src.date().isoformat() if isinstance(src, datetime.datetime) else str(src)[:10]
        d_ok = str(r[SD]) == srcd
        h_ok = bool(r[WH]) and 'neither' not in str(r[WH])
        if d_ok and h_ok:   both += 1
        elif d_ok:          date_only += 1
        elif h_ok:          host_only += 1
        else:               neither += 1
    return both, date_only, host_only, neither, blank

print(f"{'file':26s} {'BOTH':>6} {'date':>6} {'host':>6} {'none':>6} {'blank':>6}")
print("-" * 62)
for f in ('meeting_id_results.xlsx','results_exact.xlsx','v2_window1.xlsx','v2_exact.xlsx'):
    try:
        b,d,ho,n,bl = analyse(f)
        print(f"{f:26s} {b:6d} {d:6d} {ho:6d} {n:6d} {bl:6d}")
    except FileNotFoundError:
        print(f"{f:26s}  (not found)")
print("\nBOTH = S3 date matches the source date AND the host is the named")
print("       trainer or proxy. Highest column wins.")
